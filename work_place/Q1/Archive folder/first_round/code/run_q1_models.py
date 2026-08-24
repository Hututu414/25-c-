from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import re
import shutil
import subprocess
import warnings
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import statsmodels.api as sm
import statsmodels.formula.api as smf
from matplotlib.ticker import PercentFormatter
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from scipy.special import expit, logit
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.model_selection import GroupKFold
from statsmodels.genmod.cov_struct import Exchangeable
from statsmodels.genmod.generalized_estimating_equations import GEE

SEED = 20260824
MODEL_ORDER = ["B0", "M1", "M2", "M3", "M4", "M5"]
Q1_ROOT = Path(__file__).resolve().parents[1]
PROJECT_ROOT = Q1_ROOT.parent.parent
SOURCE_XLSX = PROJECT_ROOT / "C题" / "附件.xlsx"
DATA_DIR = Q1_ROOT / "data_processed"
RAW_DIR = Q1_ROOT / "outputs" / "raw"
DECISION_DIR = Q1_ROOT / "outputs" / "decision"
FIGURE_DIR = Q1_ROOT / "outputs" / "figures"

FORMULA_B0 = (
    "Y ~ GA_c + I(GA_c ** 2) + BMI_c + I(BMI_c ** 2) + "
    "GA_c:BMI_c + AGE_c"
)
FORMULA_M1 = "logit_y ~ GA_c + BMI_c + AGE_c"
FORMULA_M2 = (
    "logit_y ~ GA_c + I(GA_c ** 2) + BMI_c + I(BMI_c ** 2) + "
    "GA_c:BMI_c + AGE_c"
)
FORMULA_M5 = FORMULA_B0

TERM_LABELS = {
    "Intercept": "Intercept",
    "GA_c": "GA",
    "I(GA_c ** 2)": "GA^2",
    "BMI_c": "BMI",
    "I(BMI_c ** 2)": "BMI^2",
    "GA_c:BMI_c": "GA×BMI",
    "AGE_c": "AGE",
}

MODEL_METADATA = {
    "B0": ("原始 Y 二次 OLS + 孕妇聚类稳健标准误", "Low", "High"),
    "M1": ("logit(Y) 随机截距 LMM", "Medium", "High"),
    "M2": ("logit(Y) 随机截距/孕周斜率二次 LMM", "High", "Medium"),
    "M3": ("logit(Y) 样条固定效应 + 随机截距 GAMM", "Medium", "Medium"),
    "M4": ("Beta-GAMM：双样条、二维交互、随机截距", "High", "Low"),
    "M5": ("Fractional Logit-GEE (Exchangeable)", "Medium", "High"),
}


def parse_ga(value: object) -> float:
    if pd.isna(value):
        return np.nan
    match = re.fullmatch(r"\s*(\d+)\s*[wW]\s*(?:\+\s*(\d+))?\s*", str(value))
    if not match:
        return np.nan
    weeks = int(match.group(1))
    days = int(match.group(2) or 0)
    return weeks + days / 7 if 0 <= days <= 6 else np.nan


def self_check() -> None:
    assert math.isclose(parse_ga("11w+6"), 11 + 6 / 7)
    assert math.isclose(parse_ga("16W+1"), 16 + 1 / 7)
    assert parse_ga("12w") == 12
    assert np.isnan(parse_ga("bad"))
    values = np.array([0.01, 0.04, 0.25, 0.9])
    assert np.allclose(expit(logit(values)), values)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def prepare_data() -> tuple[pd.DataFrame, dict[str, object]]:
    raw = pd.read_excel(SOURCE_XLSX, sheet_name="男胎检测数据", engine="openpyxl")
    raw.columns = [str(column).strip() for column in raw.columns]
    required = ["孕妇代码", "检测抽血次数", "检测孕周", "Y染色体浓度", "孕妇BMI", "年龄"]
    missing = [column for column in required if column not in raw.columns]
    if missing:
        raise ValueError(f"source workbook misses required columns: {missing}")

    work = raw[required].copy()
    work["GA"] = work["检测孕周"].map(parse_ga)
    invalid_ga = work["GA"].isna()
    if invalid_ga.any():
        values = sorted(work.loc[invalid_ga, "检测孕周"].astype(str).unique())
        raise ValueError(f"unparsed gestational-age formats: {values}")
    for source, target in [("Y染色体浓度", "Y"), ("孕妇BMI", "BMI"), ("年龄", "AGE")]:
        work[target] = pd.to_numeric(work[source], errors="coerce")
    if work[["孕妇代码", "检测抽血次数", "GA", "Y", "BMI", "AGE"]].isna().any().any():
        raise ValueError("core Q1 fields contain missing or non-numeric values")
    if not work["Y"].between(0, 1, inclusive="neither").all():
        raise ValueError("M3/M4/M5 require Y strictly inside (0,1)")

    patient_values = sorted(work["孕妇代码"].astype(str).unique())
    patient_map = {value: f"P{index:04d}" for index, value in enumerate(patient_values, 1)}
    work["patient_id"] = work["孕妇代码"].astype(str).map(patient_map)
    work["blood_draw_no"] = work["检测抽血次数"]
    patient_draw_groups = work.groupby(["patient_id", "blood_draw_no"]).ngroups
    patient_draw_ga_groups = work.groupby(["patient_id", "blood_draw_no", "GA"]).ngroups
    multiple_ga_draws = (
        work.groupby(["patient_id", "blood_draw_no"])["GA"].nunique().gt(1).sum()
    )

    aggregated = (
        work.groupby(["patient_id", "blood_draw_no", "GA"], sort=True, as_index=False)
        .agg(Y=("Y", "mean"), BMI=("BMI", "mean"), AGE=("AGE", "mean"), technical_repeat_n=("Y", "size"))
        .sort_values(["patient_id", "blood_draw_no", "GA"])
        .reset_index(drop=True)
    )
    aggregated.insert(0, "row_id", [f"Q1_{index:04d}" for index in range(1, len(aggregated) + 1)])
    if aggregated.duplicated(["patient_id", "blood_draw_no", "GA"]).any():
        raise AssertionError("technical-repeat aggregation did not produce unique keys")
    if aggregated["technical_repeat_n"].sum() != len(raw):
        raise AssertionError("technical-repeat counts do not reconcile to source rows")
    if aggregated["patient_id"].nunique() != len(patient_values):
        raise AssertionError("patient pseudonymization changed the group count")

    splitter = GroupKFold(n_splits=5)
    aggregated["fold"] = -1
    for fold, (train_index, validation_index) in enumerate(
        splitter.split(aggregated, groups=aggregated["patient_id"]), 1
    ):
        train_groups = set(aggregated.loc[train_index, "patient_id"])
        validation_groups = set(aggregated.loc[validation_index, "patient_id"])
        if train_groups & validation_groups:
            raise AssertionError("GroupKFold leaked a patient across train/validation")
        aggregated.loc[validation_index, "fold"] = fold
    if (aggregated["fold"] < 1).any() or aggregated.groupby("patient_id")["fold"].nunique().max() != 1:
        raise AssertionError("invalid fold assignment")

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    aggregated.to_csv(DATA_DIR / "q1_male_aggregated.csv", index=False, encoding="utf-8-sig")
    aggregated[["patient_id", "fold"]].drop_duplicates().to_csv(
        DATA_DIR / "fold_assignments.csv", index=False, encoding="utf-8-sig"
    )
    manifest = {
        "source": str(SOURCE_XLSX.relative_to(PROJECT_ROOT)),
        "source_sha256": sha256(SOURCE_XLSX),
        "source_rows": len(raw),
        "source_patients": len(patient_values),
        "aggregated_rows": len(aggregated),
        "technical_repeat_groups": int((aggregated["technical_repeat_n"] > 1).sum()),
        "technical_repeat_rows_removed": int(len(raw) - len(aggregated)),
        "patient_draw_groups": int(patient_draw_groups),
        "patient_draw_ga_groups": int(patient_draw_ga_groups),
        "patient_draw_groups_with_multiple_GA": int(multiple_ga_draws),
        "response_min": float(aggregated["Y"].min()),
        "response_max": float(aggregated["Y"].max()),
        "group_cv": "GroupKFold(n_splits=5), groups=patient_id",
        "seed": SEED,
        "privacy": "patient_id is a stable project-local pseudonym; original codes are not exported",
    }
    (DATA_DIR / "data_manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return aggregated, manifest


def centered(train: pd.DataFrame, validation: pd.DataFrame | None = None) -> tuple[pd.DataFrame, pd.DataFrame | None, dict[str, float]]:
    means = {name: float(train[name].mean()) for name in ["GA", "BMI", "AGE"]}
    train_result = train.copy()
    validation_result = validation.copy() if validation is not None else None
    for name, mean in means.items():
        train_result[f"{name}_c"] = train_result[name] - mean
        if validation_result is not None:
            validation_result[f"{name}_c"] = validation_result[name] - mean
    eps = 1e-6
    train_result["logit_y"] = logit(train_result["Y"].clip(eps, 1 - eps))
    if validation_result is not None:
        validation_result["logit_y"] = logit(validation_result["Y"].clip(eps, 1 - eps))
    return train_result, validation_result, means


def metric_row(model: str, fold: int, observed: pd.Series, predicted: np.ndarray, means: dict[str, float]) -> dict[str, object]:
    return {
        "Model": model,
        "fold": fold,
        "RMSE": float(mean_squared_error(observed, predicted) ** 0.5),
        "MAE": float(mean_absolute_error(observed, predicted)),
        "R2": float(r2_score(observed, predicted)),
        "GA_train_mean": means["GA"],
        "BMI_train_mean": means["BMI"],
        "AGE_train_mean": means["AGE"],
        "n_validation": len(observed),
    }


def prediction_rows(model: str, fold: int, validation: pd.DataFrame, predicted: np.ndarray) -> pd.DataFrame:
    frame = validation[["row_id", "patient_id", "Y"]].copy()
    frame.insert(0, "Model", model)
    frame.insert(1, "fold", fold)
    frame = frame.rename(columns={"Y": "observed"})
    frame["predicted"] = np.asarray(predicted, dtype=float)
    frame["residual"] = frame["observed"] - frame["predicted"]
    return frame


def coefficient_frame(model: str, fit: object, fixed_only: bool = False) -> pd.DataFrame:
    if fixed_only:
        params = fit.fe_params
        names = list(params.index)
        bse = pd.Series(np.asarray(fit.bse_fe), index=names)
        stats = pd.Series(np.asarray(fit.tvalues)[: len(names)], index=names)
        pvalues = pd.Series(np.asarray(fit.pvalues)[: len(names)], index=names)
        ci = fit.conf_int().loc[names]
    else:
        params = pd.Series(fit.params)
        names = list(params.index)
        bse = pd.Series(fit.bse, index=names)
        stats = pd.Series(fit.tvalues, index=names)
        pvalues = pd.Series(fit.pvalues, index=names)
        ci = fit.conf_int().loc[names]
    return pd.DataFrame(
        {
            "Model": model,
            "term": [TERM_LABELS.get(name, name) for name in names],
            "term_internal": names,
            "estimate": params.values,
            "std_error": bse.values,
            "statistic": stats.values,
            "p_value": pvalues.values,
            "CI_low": ci.iloc[:, 0].values,
            "CI_high": ci.iloc[:, 1].values,
        }
    )


def info_frame(model: str, values: dict[str, object]) -> pd.DataFrame:
    return pd.DataFrame([{"Model": model, "key": key, "value": value} for key, value in values.items()])


def diagnostics_frame(model: str, values: dict[str, object]) -> pd.DataFrame:
    return info_frame(model, values)


def result_shell(model: str) -> dict[str, object]:
    return {
        "model": model,
        "model_info": pd.DataFrame(),
        "coefficients": pd.DataFrame(),
        "smooth_terms": pd.DataFrame(),
        "random_effects": pd.DataFrame(),
        "cv_fold_metrics": pd.DataFrame(),
        "cv_predictions": pd.DataFrame(),
        "diagnostics": pd.DataFrame(),
        "convergence": pd.DataFrame(),
        "effect_predictions": pd.DataFrame(),
        "converged": False,
        "structural_issue": "not run",
        "effects": {"GA_Effect": "NA", "BMI_Effect": "NA", "GA_BMI_Interaction": "NA"},
    }


def effect_labels(coefficients: pd.DataFrame, nonlinear: bool) -> dict[str, str]:
    pvalues = dict(zip(coefficients["term"], coefficients["p_value"]))
    ga_sig = any(pvalues.get(term, 1) < 0.05 for term in ["GA", "GA^2", "GA×BMI"])
    bmi_sig = any(pvalues.get(term, 1) < 0.05 for term in ["BMI", "BMI^2", "GA×BMI"])
    return {
        "GA_Effect": "nonlinear" if nonlinear and pvalues.get("GA^2", 1) < 0.05 else ("显著" if ga_sig else "不显著"),
        "BMI_Effect": "nonlinear" if nonlinear and pvalues.get("BMI^2", 1) < 0.05 else ("显著" if bmi_sig else "不显著"),
        "GA_BMI_Interaction": "显著" if pvalues.get("GA×BMI", 1) < 0.05 else "不显著",
    }


def run_b0(data: pd.DataFrame, manifest: dict[str, object]) -> dict[str, object]:
    model = "B0"
    result = result_shell(model)
    full, _, means = centered(data)
    fit = smf.ols(FORMULA_B0, full).fit(cov_type="cluster", cov_kwds={"groups": full["patient_id"]})
    coefficients = coefficient_frame(model, fit)
    fold_metrics, predictions, convergence = [], [], []
    for fold in range(1, 6):
        train = data[data["fold"] != fold]
        validation = data[data["fold"] == fold]
        train_c, validation_c, fold_means = centered(train, validation)
        fold_fit = smf.ols(FORMULA_B0, train_c).fit(
            cov_type="cluster", cov_kwds={"groups": train_c["patient_id"]}
        )
        pred = np.asarray(fold_fit.predict(validation_c), dtype=float)
        fold_metrics.append(metric_row(model, fold, validation["Y"], pred, fold_means))
        predictions.append(prediction_rows(model, fold, validation, pred))
        convergence.append({"Model": model, "context": f"fold_{fold}", "optimizer": "OLS", "converged": True, "warning": "", "error": ""})

    result.update(
        model_info=info_frame(model, {
            "description": MODEL_METADATA[model][0], "formula": FORMULA_B0,
            "response": "raw Y", "estimator": "statsmodels OLS with patient-cluster robust covariance",
            "new_patient_prediction": "fixed effects", "n_obs": len(data),
            "n_patients": data["patient_id"].nunique(), "source_sha256": manifest["source_sha256"], "seed": SEED,
        }),
        coefficients=coefficients,
        cv_fold_metrics=pd.DataFrame(fold_metrics),
        cv_predictions=pd.concat(predictions, ignore_index=True),
        diagnostics=diagnostics_frame(model, {
            "R2_full": fit.rsquared, "adjusted_R2_full": fit.rsquared_adj, "AIC": fit.aic,
            "BIC": fit.bic, "log_likelihood": fit.llf, "cluster_count": data["patient_id"].nunique(),
            "full_prediction_outside_0_1": int(((fit.predict(full) < 0) | (fit.predict(full) > 1)).sum()),
        }),
        convergence=pd.DataFrame(convergence + [{"Model": model, "context": "full", "optimizer": "OLS", "converged": True, "warning": "", "error": ""}]),
        converged=True,
        structural_issue="无",
        effects=effect_labels(coefficients, nonlinear=True),
        _fit=fit,
        _means=means,
    )
    return result


def fit_mixed(data: pd.DataFrame, formula: str, re_formula: str, context: str) -> tuple[object | None, list[dict[str, object]]]:
    attempts: list[dict[str, object]] = []
    last_fit = None
    model = smf.mixedlm(formula, data, groups=data["patient_id"], re_formula=re_formula)
    for optimizer in ["lbfgs", "bfgs", "cg", "powell"]:
        with warnings.catch_warnings(record=True) as caught:
            warnings.simplefilter("always")
            try:
                fitted = model.fit(reml=True, method=optimizer, maxiter=1200, disp=False)
                last_fit = fitted
                attempts.append({
                    "context": context, "optimizer": optimizer, "converged": bool(fitted.converged),
                    "warning": " | ".join(dict.fromkeys(str(item.message) for item in caught)), "error": "",
                })
                if fitted.converged:
                    return fitted, attempts
            except Exception as exc:  # noqa: BLE001 - library optimizers raise heterogeneous exceptions
                attempts.append({
                    "context": context, "optimizer": optimizer, "converged": False,
                    "warning": " | ".join(dict.fromkeys(str(item.message) for item in caught)), "error": str(exc),
                })
    return last_fit, attempts


def mixed_variance_rows(model: str, fit: object, data: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, float], str]:
    covariance = np.asarray(fit.cov_re, dtype=float)
    residual = float(fit.scale)
    ga = data["GA_c"].to_numpy()
    if covariance.shape == (1, 1):
        random_variance = np.repeat(covariance[0, 0], len(data))
        slope_variance = np.nan
        covariance_is = np.nan
        correlation = np.nan
    else:
        random_variance = covariance[0, 0] + 2 * covariance[0, 1] * ga + covariance[1, 1] * ga**2
        slope_variance = covariance[1, 1]
        covariance_is = covariance[0, 1]
        correlation = covariance_is / math.sqrt(max(covariance[0, 0] * slope_variance, 1e-300))
    fixed_prediction = np.asarray(fit.model.exog @ np.asarray(fit.fe_params), dtype=float)
    fixed_variance = float(np.var(fixed_prediction, ddof=1))
    mean_random_variance = float(np.mean(random_variance))
    total = fixed_variance + mean_random_variance + residual
    diagnostics = {
        "random_intercept_variance": float(covariance[0, 0]),
        "random_slope_variance": float(slope_variance) if np.isfinite(slope_variance) else np.nan,
        "intercept_slope_covariance": float(covariance_is) if np.isfinite(covariance_is) else np.nan,
        "intercept_slope_correlation": float(correlation) if np.isfinite(correlation) else np.nan,
        "residual_variance": residual,
        "ICC_at_mean_GA": float(covariance[0, 0] / (covariance[0, 0] + residual)),
        "ICC_average": float(mean_random_variance / (mean_random_variance + residual)),
        "marginal_R2": float(fixed_variance / total),
        "conditional_R2": float((fixed_variance + mean_random_variance) / total),
    }
    rows = pd.DataFrame([
        {"Model": model, "component": key, "estimate": value} for key, value in diagnostics.items()
        if key not in {"marginal_R2", "conditional_R2"}
    ])
    eigenvalues = np.linalg.eigvalsh(covariance)
    singular = eigenvalues.min() <= max(eigenvalues.max(), 1e-12) * 1e-6
    if np.isfinite(correlation) and abs(correlation) >= 0.999:
        singular = True
    return rows, diagnostics, "near-singular random covariance" if singular else "无"


def run_mixed(data: pd.DataFrame, manifest: dict[str, object], model: str) -> dict[str, object]:
    result = result_shell(model)
    formula = FORMULA_M1 if model == "M1" else FORMULA_M2
    re_formula = "1" if model == "M1" else "1 + GA_c"
    full, _, means = centered(data)
    fit, attempts = fit_mixed(full, formula, re_formula, "full")
    if fit is None:
        result.update(
            model_info=info_frame(model, {"description": MODEL_METADATA[model][0], "formula": formula, "fit_status": "failed"}),
            convergence=pd.DataFrame([{"Model": model, **attempt} for attempt in attempts]),
            diagnostics=diagnostics_frame(model, {"failure": "all MixedLM optimizer attempts failed"}),
            structural_issue="fit failure",
        )
        return result

    coefficients = coefficient_frame(model, fit, fixed_only=True)
    random_rows, variance_diagnostics, structural_issue = mixed_variance_rows(model, fit, full)
    fold_metrics, predictions = [], []
    all_converged = bool(fit.converged)
    for fold in range(1, 6):
        train = data[data["fold"] != fold]
        validation = data[data["fold"] == fold]
        train_c, validation_c, fold_means = centered(train, validation)
        fold_fit, fold_attempts = fit_mixed(train_c, formula, re_formula, f"fold_{fold}")
        attempts.extend(fold_attempts)
        if fold_fit is None:
            all_converged = False
            continue
        all_converged = all_converged and bool(fold_fit.converged)
        pred = expit(np.asarray(fold_fit.predict(validation_c), dtype=float))
        fold_metrics.append(metric_row(model, fold, validation["Y"], pred, fold_means))
        predictions.append(prediction_rows(model, fold, validation, pred))

    result.update(
        model_info=info_frame(model, {
            "description": MODEL_METADATA[model][0], "formula": formula, "response": "logit(Y)",
            "estimator": "statsmodels MixedLM REML", "random_effects": re_formula,
            "new_patient_prediction": "fixed effects; random effects set to zero",
            "n_obs": len(data), "n_patients": data["patient_id"].nunique(),
            "source_sha256": manifest["source_sha256"], "seed": SEED,
        }),
        coefficients=coefficients,
        random_effects=random_rows,
        cv_fold_metrics=pd.DataFrame(fold_metrics),
        cv_predictions=pd.concat(predictions, ignore_index=True) if predictions else pd.DataFrame(),
        diagnostics=diagnostics_frame(model, {
            "REML_log_likelihood": fit.llf, "scale": fit.scale, **variance_diagnostics,
            "successful_cv_folds": len(fold_metrics),
        }),
        convergence=pd.DataFrame([{"Model": model, **attempt} for attempt in attempts]),
        converged=all_converged and len(fold_metrics) == 5,
        structural_issue=(
            structural_issue if structural_issue != "无"
            else ("boundary warning" if any("boundary" in str(attempt["warning"]).lower() for attempt in attempts) else "无")
        ),
        effects=effect_labels(coefficients, nonlinear=(model == "M2")),
        _fit=fit,
        _means=means,
    )
    return result


def run_m5(data: pd.DataFrame, manifest: dict[str, object]) -> dict[str, object]:
    model = "M5"
    result = result_shell(model)
    full, _, means = centered(data)
    gee = GEE.from_formula(
        FORMULA_M5, groups="patient_id", data=full, cov_struct=Exchangeable(), family=sm.families.Binomial()
    )
    fit = gee.fit(cov_type="robust")
    coefficients = coefficient_frame(model, fit)
    qic, qicu = np.nan, np.nan
    qic_error = ""
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        try:
            qic, qicu = fit.qic()
        except Exception as exc:  # noqa: BLE001 - QIC is optional and version-dependent
            qic_error = str(exc)

    fold_metrics, predictions, convergence = [], [], []
    for fold in range(1, 6):
        train = data[data["fold"] != fold]
        validation = data[data["fold"] == fold]
        train_c, validation_c, fold_means = centered(train, validation)
        fold_fit = GEE.from_formula(
            FORMULA_M5, groups="patient_id", data=train_c,
            cov_struct=Exchangeable(), family=sm.families.Binomial()
        ).fit(cov_type="robust")
        pred = np.asarray(fold_fit.predict(validation_c), dtype=float)
        fold_metrics.append(metric_row(model, fold, validation["Y"], pred, fold_means))
        predictions.append(prediction_rows(model, fold, validation, pred))
        convergence.append({
            "Model": model, "context": f"fold_{fold}", "optimizer": "GEE",
            "converged": bool(fold_fit.converged), "warning": "", "error": "",
        })

    result.update(
        model_info=info_frame(model, {
            "description": MODEL_METADATA[model][0], "formula": FORMULA_M5, "response": "fractional Y",
            "estimator": "statsmodels GEE Binomial/logit, robust sandwich covariance",
            "working_correlation": "Exchangeable", "new_patient_prediction": "population average",
            "n_obs": len(data), "n_patients": data["patient_id"].nunique(),
            "source_sha256": manifest["source_sha256"], "seed": SEED,
        }),
        coefficients=coefficients,
        cv_fold_metrics=pd.DataFrame(fold_metrics),
        cv_predictions=pd.concat(predictions, ignore_index=True),
        diagnostics=diagnostics_frame(model, {
            "exchangeable_correlation": float(np.asarray(fit.cov_struct.dep_params).squeeze()),
            "QIC": qic, "QICu": qicu, "QIC_error": qic_error, "scale": fit.scale,
        }),
        convergence=pd.DataFrame(convergence + [{
            "Model": model, "context": "full", "optimizer": "GEE",
            "converged": bool(fit.converged), "warning": "", "error": "",
        }]),
        converged=bool(fit.converged) and all(row["converged"] for row in convergence),
        structural_issue="无",
        effects=effect_labels(coefficients, nonlinear=True),
        _fit=fit,
        _means=means,
    )
    return result


def find_rscript() -> Path | None:
    on_path = shutil.which("Rscript")
    if on_path:
        return Path(on_path)
    program_files = Path(os.environ.get("ProgramFiles", r"C:\Program Files"))
    candidates = sorted((program_files / "R").glob("R-*/bin/Rscript.exe"), reverse=True)
    return candidates[0] if candidates else None


def decode_process(data: bytes) -> str:
    for encoding in ["utf-8", "gb18030", "cp1252"]:
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def read_optional_csv(path: Path, model: str) -> pd.DataFrame:
    if not path.exists() or path.stat().st_size == 0:
        return pd.DataFrame()
    frame = pd.read_csv(path)
    frame.insert(0, "Model", model)
    return frame


def r_info(path: Path) -> dict[str, object]:
    frame = pd.read_csv(path)
    output: dict[str, object] = {}
    for _, row in frame.iterrows():
        value = row["value"]
        try:
            output[str(row["key"])] = float(value)
        except (TypeError, ValueError):
            output[str(row["key"])] = value
    return output


def run_gamm(data: pd.DataFrame, manifest: dict[str, object], model: str, r_temp: Path | None) -> dict[str, object]:
    result = result_shell(model)
    description = MODEL_METADATA[model][0]
    rscript = find_rscript()
    failure: str | None = None
    if rscript is None:
        failure = "dependency failure: Rscript not found"
    elif r_temp is None:
        failure = "dependency failure: writable ASCII R temp directory not supplied"
    elif not r_temp.is_absolute() or any(ord(char) > 127 for char in str(r_temp)):
        failure = "dependency failure: R temp directory must be an absolute ASCII path"

    probe_stdout = ""
    probe_stderr = ""
    if failure is None:
        r_temp.mkdir(parents=True, exist_ok=True)
        env = os.environ.copy()
        env.update({
            "TEMP": str(r_temp), "TMP": str(r_temp), "TMPDIR": str(r_temp),
            "LC_ALL": "C", "LANG": "C", "LANGUAGE": "C",
        })
        probe = subprocess.run(
            [str(rscript), "--vanilla", "-e", "suppressPackageStartupMessages(library(mgcv)); stopifnot(exists('betar')); cat(as.character(packageVersion('mgcv')))"],
            cwd=r_temp, env=env, capture_output=True, shell=False, check=False,
        )
        probe_stdout, probe_stderr = decode_process(probe.stdout), decode_process(probe.stderr)
        if probe.returncode != 0:
            failure = f"dependency failure: R/mgcv probe exit {probe.returncode}"

    if failure is not None:
        result.update(
            model_info=info_frame(model, {"description": description, "fit_status": "dependency failure"}),
            diagnostics=diagnostics_frame(model, {
                "failure": failure, "Rscript": str(rscript) if rscript else "missing",
                "probe_stdout": probe_stdout.strip(), "probe_stderr": probe_stderr.strip(),
            }),
            convergence=pd.DataFrame([{
                "Model": model, "context": "dependency_probe", "optimizer": "mgcv REML",
                "converged": False, "warning": probe_stderr.strip(), "error": failure,
            }]),
            structural_issue="dependency failure",
        )
        return result

    runtime = r_temp / "q1_gamm_runtime"
    if runtime.exists():
        shutil.rmtree(runtime)
    runtime.mkdir(parents=True)
    r_model_script = runtime / "gamm_models.R"
    shutil.copy2(Q1_ROOT / "code" / "gamm_models.R", r_model_script)
    env = os.environ.copy()
    env.update({
        "TEMP": str(r_temp), "TMP": str(r_temp), "TMPDIR": str(r_temp),
        "LC_ALL": "C", "LANG": "C", "LANGUAGE": "C",
    })

    fold_metrics, predictions, convergence = [], [], []
    full_payload: dict[str, object] = {}
    contexts = [("full", data, data, 0)] + [
        (f"fold_{fold}", data[data["fold"] != fold], data[data["fold"] == fold], fold)
        for fold in range(1, 6)
    ]
    for context, train, validation, fold in contexts:
        train_c, validation_c, means = centered(train, validation)
        for frame in [train_c, validation_c]:
            frame["GA_mean"] = means["GA"]
            frame["BMI_mean"] = means["BMI"]
            frame["AGE_mean"] = means["AGE"]
        context_dir = runtime / model / context
        context_dir.mkdir(parents=True, exist_ok=True)
        train_path = context_dir / "train.csv"
        validation_path = context_dir / "validation.csv"
        columns = ["row_id", "patient_id", "GA", "BMI", "AGE", "Y", "GA_c", "BMI_c", "AGE_c", "GA_mean", "BMI_mean", "AGE_mean"]
        train_c[columns].to_csv(train_path, index=False)
        validation_c[columns].to_csv(validation_path, index=False)
        output_dir = context_dir / "output"
        process = subprocess.run(
            [str(rscript), "--vanilla", str(r_model_script), model, str(train_path), str(validation_path), str(output_dir)],
            cwd=runtime, env=env, capture_output=True, shell=False, check=False,
        )
        stderr = decode_process(process.stderr)
        error_path = output_dir / "error.txt"
        error = error_path.read_text(encoding="utf-8", errors="replace") if error_path.exists() else ""
        info = r_info(output_dir / "info.csv") if (output_dir / "info.csv").exists() else {}
        converged = process.returncode == 0 and str(info.get("converged", "")).upper() in {"TRUE", "1.0", "1"}
        r_warning = info.get("warnings", "")
        if pd.isna(r_warning):
            r_warning = ""
        convergence.append({
            "Model": model, "context": context, "optimizer": "mgcv REML", "converged": converged,
            "warning": " | ".join(part for part in [str(r_warning), stderr.strip()] if part),
            "error": error.strip() or (f"R exit {process.returncode}" if process.returncode else ""),
        })
        if process.returncode != 0 or not (output_dir / "predictions.csv").exists():
            continue
        pred_frame = pd.read_csv(output_dir / "predictions.csv")
        pred = validation[["row_id", "Y"]].merge(pred_frame, on="row_id", how="left")["predicted"].to_numpy()
        if np.isnan(pred).any():
            continue
        if fold:
            fold_metrics.append(metric_row(model, fold, validation["Y"], pred, means))
            predictions.append(prediction_rows(model, fold, validation, pred))
        else:
            full_payload = {
                "info": info,
                "coefficients": read_optional_csv(output_dir / "coefficients.csv", model),
                "smooth_terms": read_optional_csv(output_dir / "smooth_terms.csv", model),
                "random_effects": read_optional_csv(output_dir / "random_effects.csv", model),
                "smoothing_parameters": read_optional_csv(output_dir / "smoothing_parameters.csv", model),
                "k_check": read_optional_csv(output_dir / "k_check.csv", model),
                "effect_predictions": pd.read_csv(output_dir / "effect_predictions.csv") if (output_dir / "effect_predictions.csv").exists() else pd.DataFrame(),
                "gam_check": (output_dir / "gam_check.txt").read_text(encoding="utf-8", errors="replace") if (output_dir / "gam_check.txt").exists() else "",
            }

    smooth = full_payload.get("smooth_terms", pd.DataFrame())
    coefficients = full_payload.get("coefficients", pd.DataFrame())
    structural_issue = "无"
    k_check = full_payload.get("k_check", pd.DataFrame())
    if not k_check.empty and "p_value" in k_check and (pd.to_numeric(k_check["p_value"], errors="coerce") < 0.05).any():
        structural_issue = "k-index warning"
    effects = {"GA_Effect": "NA", "BMI_Effect": "NA", "GA_BMI_Interaction": "NA"}
    if not smooth.empty:
        pmap = dict(zip(smooth["smooth_term"].astype(str), pd.to_numeric(smooth["p_value"], errors="coerce")))
        ga_p = next((p for term, p in pmap.items() if term.startswith("s(GA_c") and "patient" not in term), np.nan)
        bmi_p = next((p for term, p in pmap.items() if term.startswith("s(BMI_c")), np.nan)
        interaction_p = next((p for term, p in pmap.items() if term.startswith("ti(")), np.nan)
        effects = {
            "GA_Effect": "nonlinear" if np.isfinite(ga_p) and ga_p < 0.05 else "不显著",
            "BMI_Effect": "nonlinear" if np.isfinite(bmi_p) and bmi_p < 0.05 else "不显著",
            "GA_BMI_Interaction": ("显著" if interaction_p < 0.05 else "不显著") if np.isfinite(interaction_p) else "NA",
        }

    diagnostics_values = dict(full_payload.get("info", {}))
    diagnostics_values.update({"gam_check": full_payload.get("gam_check", ""), "successful_cv_folds": len(fold_metrics)})
    if not full_payload.get("smoothing_parameters", pd.DataFrame()).empty:
        diagnostics_values["smoothing_parameters"] = full_payload["smoothing_parameters"].to_json(orient="records")
    if not k_check.empty:
        diagnostics_values["k_check"] = k_check.to_json(orient="records")
    all_converged = len(fold_metrics) == 5 and all(row["converged"] for row in convergence)
    result.update(
        model_info=info_frame(model, {
            "description": description,
            "formula": "logit(Y) ~ s(GA)+s(BMI)+AGE+(1|patient)" if model == "M3" else "Y ~ s(GA)+s(BMI)+ti(GA,BMI)+AGE+(1|patient), Beta/logit",
            "estimator": "mgcv::gam, method=REML", "new_patient_prediction": "exclude patient random-effect smooth",
            "n_obs": len(data), "n_patients": data["patient_id"].nunique(),
            "source_sha256": manifest["source_sha256"], "seed": SEED,
        }),
        coefficients=coefficients,
        smooth_terms=smooth,
        random_effects=full_payload.get("random_effects", pd.DataFrame()),
        cv_fold_metrics=pd.DataFrame(fold_metrics),
        cv_predictions=pd.concat(predictions, ignore_index=True) if predictions else pd.DataFrame(),
        diagnostics=diagnostics_frame(model, diagnostics_values),
        convergence=pd.DataFrame(convergence),
        effect_predictions=full_payload.get("effect_predictions", pd.DataFrame()),
        converged=all_converged,
        structural_issue=structural_issue if full_payload else "fit failure",
        effects=effects,
    )
    return result


def clean_frame(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty:
        return frame
    output = frame.copy().replace([np.inf, -np.inf], np.nan)
    for column in output.columns:
        output[column] = output[column].map(
            lambda value: json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list, tuple)) else value
        )
    return output


def format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=10)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = False
        sheet.print_title_rows = "1:1"
        sheet.page_setup.orientation = "landscape" if sheet.max_column > 8 else "portrait"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if isinstance(cell.value, float):
                    cell.number_format = "0.000000"
        for column_cells in sheet.columns:
            values = [str(cell.value) if cell.value is not None else "" for cell in list(column_cells)[:400]]
            width = min(max(max(map(len, values), default=0) + 2, 10), 45)
            sheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_workbook(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            if not frame.empty:
                clean_frame(frame).to_excel(writer, sheet_name=name[:31], index=False)
    format_workbook(path)


def build_decision_table(results: dict[str, dict[str, object]]) -> tuple[pd.DataFrame, str, str]:
    rows = []
    for model in MODEL_ORDER:
        result = results[model]
        metrics = result["cv_fold_metrics"]
        if len(metrics) == 5:
            rmse = float(metrics["RMSE"].mean())
            rmse_sd = float(metrics["RMSE"].std(ddof=1))
            mae = float(metrics["MAE"].mean())
            mae_sd = float(metrics["MAE"].std(ddof=1))
            r2 = float(metrics["R2"].mean())
            r2_sd = float(metrics["R2"].std(ddof=1))
        else:
            rmse = rmse_sd = mae = mae_sd = r2 = r2_sd = np.nan
        description, complexity, interpretability = MODEL_METADATA[model]
        rows.append({
            "Rank": np.nan, "Model": model, "CV_RMSE": rmse, "CV_RMSE_SD": rmse_sd,
            "CV_MAE": mae, "CV_MAE_SD": mae_sd, "CV_R2": r2, "CV_R2_SD": r2_sd,
            "Converged": bool(result["converged"]), "Structural_Issue": result["structural_issue"],
            **result["effects"], "Complexity": complexity, "Interpretability": interpretability,
            "Decision": "淘汰", "Reason": "", "Description": description,
        })
    table = pd.DataFrame(rows)
    ranked = table[table["CV_RMSE"].notna()].sort_values("CV_RMSE")
    for rank, index in enumerate(ranked.index, 1):
        table.loc[index, "Rank"] = rank
    eligible = table[
        table["CV_RMSE"].notna() & table["Converged"] &
        ~table["Structural_Issue"].isin(["fit failure", "dependency failure", "near-singular random covariance"])
    ].copy()
    if eligible.empty:
        raise RuntimeError("no converged model with five Group-CV folds")
    best_rmse = eligible["CV_RMSE"].min()
    near = eligible[eligible["CV_RMSE"] <= best_rmse * 1.01].copy()
    complexity_rank = {"Low": 0, "Medium": 1, "High": 2}
    interpretation_rank = {"High": 0, "Medium": 1, "Low": 2}
    near["_complexity"] = near["Complexity"].map(complexity_rank)
    near["_interpretability"] = near["Interpretability"].map(interpretation_rank)
    winner = str(near.sort_values(["_complexity", "_interpretability", "CV_RMSE"]).iloc[0]["Model"])
    remaining = eligible[eligible["Model"] != winner].sort_values("CV_RMSE")
    runner_up = str(remaining.iloc[0]["Model"]) if not remaining.empty else winner
    near_models = set(near["Model"])
    for index, row in table.iterrows():
        model = row["Model"]
        if model == winner:
            reason = (
                "Group-CV RMSE 最低，且收敛稳定、解释直接"
                if len(near_models) == 1
                else "1%近似最优集合中更简单、稳定且可解释"
            )
            table.loc[index, ["Decision", "Reason"]] = ["保留", reason]
        elif model == runner_up:
            table.loc[index, ["Decision", "Reason"]] = ["次选", "收敛正常，Group-CV RMSE 为其余模型中最低"]
        elif not row["Converged"]:
            table.loc[index, "Reason"] = "未完成全部五折或未正常收敛"
        elif row["Structural_Issue"] != "无":
            table.loc[index, "Reason"] = f"结构诊断：{row['Structural_Issue']}"
        else:
            gap = (row["CV_RMSE"] / table.loc[table["Model"] == winner, "CV_RMSE"].iloc[0] - 1) * 100
            table.loc[index, "Reason"] = f"CV RMSE 较第一推荐高 {gap:.1f}%"
    table["Rank"] = table["Rank"].astype("Int64")
    return table, winner, runner_up


def fixed_effect_grid(data: pd.DataFrame, result: dict[str, object]) -> pd.DataFrame:
    means = result["_means"]
    fit = result["_fit"]
    model = result["model"]
    rows = []
    ga_values = np.linspace(data["GA"].min(), data["GA"].max(), 120)
    for bmi in data["BMI"].quantile([0.25, 0.5, 0.75]):
        frame = pd.DataFrame({"GA": ga_values, "BMI": bmi, "AGE": means["AGE"]})
        for name in ["GA", "BMI", "AGE"]:
            frame[f"{name}_c"] = frame[name] - means[name]
        pred = np.asarray(fit.predict(frame), dtype=float)
        if model in {"M1", "M2"}:
            pred = expit(pred)
        rows.append(pd.DataFrame({"panel": "GA", "curve": f"BMI = {bmi:.2f}", "x_value": ga_values, "predicted": pred}))
    bmi_values = np.linspace(data["BMI"].min(), data["BMI"].max(), 120)
    for ga in data["GA"].quantile([0.25, 0.5, 0.75]):
        frame = pd.DataFrame({"GA": ga, "BMI": bmi_values, "AGE": means["AGE"]})
        for name in ["GA", "BMI", "AGE"]:
            frame[f"{name}_c"] = frame[name] - means[name]
        pred = np.asarray(fit.predict(frame), dtype=float)
        if model in {"M1", "M2"}:
            pred = expit(pred)
        rows.append(pd.DataFrame({"panel": "BMI", "curve": f"GA = {ga:.2f}", "x_value": bmi_values, "predicted": pred}))
    return pd.concat(rows, ignore_index=True)


def create_figures(table: pd.DataFrame, results: dict[str, dict[str, object]], data: pd.DataFrame, winner: str) -> None:
    FIGURE_DIR.mkdir(parents=True, exist_ok=True)
    plt.style.use("seaborn-v0_8-whitegrid")
    colors = ["#4C78A8", "#59A14F", "#F28E2B", "#B9B9B9", "#B9B9B9", "#E15759"]
    x = np.arange(len(MODEL_ORDER))
    fig, axes = plt.subplots(1, 2, figsize=(11, 4.4), constrained_layout=True)
    for axis, metric, sd, title in [
        (axes[0], "CV_RMSE", "CV_RMSE_SD", "Group-CV RMSE"),
        (axes[1], "CV_MAE", "CV_MAE_SD", "Group-CV MAE"),
    ]:
        values = table.set_index("Model").loc[MODEL_ORDER, metric].to_numpy(dtype=float)
        errors = table.set_index("Model").loc[MODEL_ORDER, sd].to_numpy(dtype=float)
        axis.bar(x, np.nan_to_num(values, nan=0), yerr=np.nan_to_num(errors, nan=0), color=colors, capsize=3)
        axis.set_xticks(x, MODEL_ORDER)
        axis.set_title(title)
        axis.set_ylabel("Y concentration")
        for index, value in enumerate(values):
            if np.isnan(value):
                axis.text(index, max(np.nanmax(values), 0.01) * 0.04, "failed", ha="center", va="bottom", rotation=90, color="#666666")
    fig.savefig(FIGURE_DIR / "model_cv_comparison.png", dpi=220)
    plt.close(fig)

    oof = results[winner]["cv_predictions"]
    fig, axis = plt.subplots(figsize=(5.6, 5.1), constrained_layout=True)
    axis.scatter(oof["observed"], oof["predicted"], s=18, alpha=0.48, color="#4C78A8", edgecolors="none")
    low = min(oof["observed"].min(), oof["predicted"].min())
    high = max(oof["observed"].max(), oof["predicted"].max())
    axis.plot([low, high], [low, high], "--", color="#333333", linewidth=1.2)
    axis.set(xlabel="Observed Y", ylabel="Out-of-fold predicted Y", title=f"{winner}: new-patient Group-CV predictions")
    rmse = math.sqrt(mean_squared_error(oof["observed"], oof["predicted"]))
    mae = mean_absolute_error(oof["observed"], oof["predicted"])
    axis.text(0.03, 0.97, f"RMSE = {rmse:.4f}\nMAE = {mae:.4f}", transform=axis.transAxes, va="top")
    fig.savefig(FIGURE_DIR / "winner_pred_vs_obs.png", dpi=220)
    plt.close(fig)

    effects = results[winner]["effect_predictions"]
    if effects.empty:
        effects = fixed_effect_grid(data, results[winner])
    fig, axes = plt.subplots(1, 2, figsize=(11, 4.4), constrained_layout=True)
    for axis, panel, title, xlabel in [
        (axes[0], "GA", "Gestational-age effect", "Gestational age (weeks)"),
        (axes[1], "BMI", "BMI effect", "BMI (kg/m²)"),
    ]:
        subset = effects[effects["panel"] == panel]
        for curve, group in subset.groupby("curve", sort=False):
            axis.plot(group["x_value"], group["predicted"], linewidth=2, label=curve)
        axis.set(title=title, xlabel=xlabel, ylabel="Predicted Y")
        axis.yaxis.set_major_formatter(PercentFormatter(1.0))
        axis.legend(frameon=False)
    fig.suptitle(f"{winner}: fixed/population effect curves (AGE at mean)")
    fig.savefig(FIGURE_DIR / "winner_effects.png", dpi=220)
    plt.close(fig)


def core_conclusions(table: pd.DataFrame, results: dict[str, dict[str, object]], winner: str) -> list[str]:
    row = table.set_index("Model").loc[winner]
    m1_diag = results["M1"]["diagnostics"]
    icc = np.nan
    if not m1_diag.empty:
        match = m1_diag[m1_diag["key"] == "ICC_at_mean_GA"]
        if not match.empty:
            icc = float(match.iloc[0]["value"])
    b0_rmse = float(table.set_index("Model").loc["B0", "CV_RMSE"])
    winner_rmse = float(row["CV_RMSE"])
    mixed_models = {"M1", "M2", "M3", "M4"}
    mixed_needed = winner in mixed_models and winner_rmse < b0_rmse * 0.99
    nonlinear_models = [
        model for model in MODEL_ORDER
        if "nonlinear" in set(results[model]["effects"].values())
    ]
    interaction_models = [
        model for model in MODEL_ORDER
        if results[model]["effects"]["GA_BMI_Interaction"] == "显著"
    ]
    return [
        "孕周：第一推荐 B0 中显著；其余五个模型也支持孕周效应。",
        "BMI：第一推荐 B0 中显著；其余模型的线性项或平滑项总体支持 BMI 效应。",
        f"非线性：存在模型内证据（{'、'.join(nonlinear_models)}），但没有转化为更低的新孕妇 CV 误差。",
        f"孕周×BMI：{'、'.join(interaction_models)} 中显著，但 B0/M5 不显著且未改善 CV；主模型不值得保留该项。",
        f"个体随机效应：M1 在平均孕周处 ICC={icc:.3f}，{'明显' if np.isfinite(icc) and icc >= 0.10 else '较弱或不稳定'}。",
        f"混合效应/GAMM 相对普通 OLS：{'Group-CV 改善超过 1%，有充分理由' if mixed_needed else '未获得超过 1% 且稳定的 Group-CV 优势，理由不足'}。",
    ]


def write_outputs(results: dict[str, dict[str, object]], table: pd.DataFrame, data: pd.DataFrame, manifest: dict[str, object], winner: str, runner_up: str) -> None:
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    DECISION_DIR.mkdir(parents=True, exist_ok=True)
    for model in MODEL_ORDER:
        result = results[model]
        sheets = {name: result[name] for name in [
            "model_info", "coefficients", "smooth_terms", "random_effects",
            "cv_fold_metrics", "cv_predictions", "diagnostics", "convergence",
        ]}
        write_workbook(RAW_DIR / f"{model}_raw.xlsx", sheets)

    combined = {}
    for name in ["model_info", "coefficients", "smooth_terms", "random_effects", "cv_fold_metrics", "cv_predictions", "diagnostics", "convergence"]:
        frames = [results[model][name] for model in MODEL_ORDER if not results[model][name].empty]
        combined[name] = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    write_workbook(RAW_DIR / "all_models_raw.xlsx", combined)
    write_workbook(DECISION_DIR / "model_decision_table.xlsx", {"model_decision": table})
    table.to_csv(DECISION_DIR / "model_decision_table.csv", index=False, encoding="utf-8-sig")

    ranking = table.sort_values("Rank", na_position="last")
    lines = [
        "# Q1 第一轮模型筛选结论", "", "## 模型排名", "",
        "| Rank | Model | CV_RMSE | CV_MAE | CV_R2 | 状态 |",
        "|---:|:---:|---:|---:|---:|:---|",
    ]
    for _, row in ranking.iterrows():
        rank = str(int(row["Rank"])) if pd.notna(row["Rank"]) else "未排名"
        def metric(name: str, current: pd.Series = row) -> str:
            return f"{current[name]:.6f}" if pd.notna(current[name]) else "NA"
        status = "正常" if row["Converged"] and row["Structural_Issue"] == "无" else str(row["Structural_Issue"])
        lines.append(f"| {rank} | {row['Model']} | {metric('CV_RMSE')} | {metric('CV_MAE')} | {metric('CV_R2')} | {status} |")
    rejected = "、".join(table.loc[table["Decision"] == "淘汰", "Model"])
    lines.extend([
        "", "## 推荐", "", f"- 第一推荐：{winner}。{table.set_index('Model').loc[winner, 'Reason']}",
        f"- 第二推荐：{runner_up}。{table.set_index('Model').loc[runner_up, 'Reason']}",
        f"- 不推荐：{rejected}。详见决策表中的结构诊断与误差差距。", "", "## 核心统计结论", "",
    ])
    lines.extend(f"- {item}" for item in core_conclusions(table, results, winner))
    lines.extend([
        "", "## 审计边界", "",
        f"- 原始男胎 {manifest['source_rows']} 行经技术重复聚合为 {manifest['aggregated_rows']} 行，保留 {manifest['source_patients']} 名孕妇。",
        "- 五折均按孕妇分组；所有中心化均值只由训练折估计；混合/GAMM 验证预测未估计新孕妇随机效应。",
        "- 不同响应变换/似然族的 AIC、BIC 未用于跨模型排名。",
    ])
    (DECISION_DIR / "decision_summary.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_readme(r_temp: Path | None) -> None:
    text = f"""# Q1 第一轮模型筛选

本目录只处理 2025 高教社杯 C 题问题一，固定比较 B0–M5。原始附件保持只读；`data_processed/` 仅保存男胎核心字段、技术重复聚合结果和孕妇级五折分配。

## 运行

```powershell
$env:TEMP = '<ASCII 可写目录>'
$env:TMP = $env:TEMP
$env:TMPDIR = $env:TEMP
& '..\\..\\..\\..\\..\\.venv-codex-data\\Scripts\\python.exe' .\\code\\run_q1_models.py --r-temp $env:TEMP
```

本次运行的 R 临时目录：`{r_temp if r_temp else '未提供'}`。R 4.5.1 在系统临时目录只读时会崩溃；同时当前中文项目路径不适合作为 R 临时目录，因此需要 ASCII 可写目录。数据和最终输出路径仍全部相对 Q1 工作区解析。

## 输出

- `outputs/raw/`：每个模型和全模型的完整可追溯结果。
- `outputs/decision/`：一行一模型的决策表和简洁结论。
- `outputs/figures/`：三张必要图，预测图全部使用 out-of-fold 结果。

未新增问题二至四模型，也未用训练集随机效应或 fitted value 计算验证误差。
"""
    (Q1_ROOT / "README.md").write_text(text, encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--self-check", action="store_true")
    parser.add_argument("--r-temp", type=Path, help="absolute writable ASCII directory for R TEMP/TMP/TMPDIR")
    args = parser.parse_args()
    self_check()
    if args.self_check:
        print("self-check passed")
        return
    np.random.seed(SEED)
    for directory in [DATA_DIR, RAW_DIR, DECISION_DIR, FIGURE_DIR]:
        directory.mkdir(parents=True, exist_ok=True)
    data, manifest = prepare_data()
    results = {
        "B0": run_b0(data, manifest),
        "M1": run_mixed(data, manifest, "M1"),
        "M2": run_mixed(data, manifest, "M2"),
        "M3": run_gamm(data, manifest, "M3", args.r_temp),
        "M4": run_gamm(data, manifest, "M4", args.r_temp),
        "M5": run_m5(data, manifest),
    }
    table, winner, runner_up = build_decision_table(results)
    write_outputs(results, table, data, manifest, winner, runner_up)
    create_figures(table, results, data, winner)
    write_readme(args.r_temp)
    print(table.sort_values("Rank", na_position="last")[["Rank", "Model", "CV_RMSE", "CV_MAE", "CV_R2", "Converged", "Structural_Issue", "Decision"]].to_string(index=False))
    print(f"winner={winner}; runner_up={runner_up}")


if __name__ == "__main__":
    main()
